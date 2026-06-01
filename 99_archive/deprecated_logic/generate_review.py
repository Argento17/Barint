"""
BSIP2 Prototype v0 — Review Dashboard Generator
Outputs: review/bsip2_master_products.csv
         review/bsip2_master_products.xlsx
         review/review_dashboard.html
"""
import json
import pathlib
import csv
import html as html_module
import datetime

TRACE_ROOT = pathlib.Path(r"C:\Bari\03_operations\bsip2\proto_v0\outputs\products")
REVIEW_ROOT = pathlib.Path(r"C:\Bari\03_operations\bsip2\proto_v0\review")

GRADE_ORDER = {"C": 0, "D": 1, "E": 2, "insufficient_data": 3}
GRADE_COLORS = {
    "C": "#2a9d8f",
    "D": "#e9c46a",
    "E": "#e76f51",
    "insufficient_data": "#aaaaaa",
}

# ── helpers ─────────────────────────────────────────────────────────────────

def load_traces():
    records = []
    for p in sorted(TRACE_ROOT.glob("*/bsip2_trace.json")):
        with open(p, encoding="utf-8") as f:
            records.append(json.load(f))
    return records


def _dominant_concern_family(t):
    """Return the concern family that drove the binding cap, or highest penalty family."""
    coord = t.get("concern_family_coordination") or {}
    binding = t.get("binding_cap")
    if binding is not None:
        for fam, data in coord.items():
            if data.get("binding_cap") == binding:
                return fam
    # No cap: family with highest coordinated_penalty
    best_fam, best_pen = None, -1
    for fam, data in coord.items():
        pen = data.get("coordinated_penalty") or 0
        if pen > best_pen:
            best_pen = pen
            best_fam = fam
    return best_fam if best_pen > 0 else "none"


def _dominant_positive_signal(t):
    """Return the most helpful positive signal (floor or best dimension)."""
    floors = t.get("floors_applied") or []
    if floors:
        f = floors[0]
        return f"floor:{f.get('rule','?')}→{f.get('floor_value','?')}"
    dim_scores = t.get("dimension_scores") or {}
    if not dim_scores:
        return "none"
    best = max(dim_scores, key=lambda k: dim_scores[k])
    return f"{best}={dim_scores[best]}"


def _structural_emptiness(t):
    ser = t.get("structural_emptiness_result") or {}
    return "yes" if ser.get("structurally_empty") else "no"


def _caps_applied_str(t):
    caps = t.get("caps_applied") or []
    if not caps:
        return ""
    return "; ".join(f"{c['rule']}→{c['cap']}" for c in caps)


def _floors_applied_str(t):
    floors = t.get("floors_applied") or []
    if not floors:
        return ""
    return "; ".join(f"{f.get('rule','?')}→{f.get('floor_value','?')}" for f in floors)


def _unresolved_flags_count(t):
    flags = t.get("unresolved_flags") or []
    return len(flags)


def build_row(t):
    ref = t.get("input_reference") or {}
    return {
        "product_id": ref.get("canonical_product_id", ""),
        "product_name": ref.get("product_name_he", ""),
        "score": t.get("final_score_estimate"),
        "grade": t.get("grade_estimate", ""),
        "category": t.get("category", ""),
        "category_confidence": t.get("category_confidence"),
        "nova_proxy": t.get("nova_proxy"),
        "confidence_score": t.get("confidence_score"),
        "evaluation_status": t.get("evaluation_status", ""),
        "unresolved_flags": _unresolved_flags_count(t),
        "caps_applied": _caps_applied_str(t),
        "floors_applied": _floors_applied_str(t),
        "dominant_concern_family": _dominant_concern_family(t),
        "dominant_positive_signal": _dominant_positive_signal(t),
        "structural_emptiness": _structural_emptiness(t),
    }


def build_all_rows(traces):
    rows = [build_row(t) for t in traces]
    rows.sort(key=lambda r: (
        GRADE_ORDER.get(r["grade"], 99),
        -(r["score"] or 0)
    ))
    return rows


# ── CSV export ───────────────────────────────────────────────────────────────

COLUMNS = [
    "product_name", "score", "grade", "category", "category_confidence",
    "nova_proxy", "confidence_score", "evaluation_status", "unresolved_flags",
    "caps_applied", "floors_applied", "dominant_concern_family",
    "dominant_positive_signal", "structural_emptiness", "product_id",
]


def export_csv(rows, out_path):
    with open(out_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=COLUMNS)
        writer.writeheader()
        for r in rows:
            writer.writerow({c: r.get(c, "") for c in COLUMNS})
    print(f"  CSV  -> {out_path}")


# ── XLSX export ──────────────────────────────────────────────────────────────

def export_xlsx(rows, out_path):
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "BSIP2 Products"

    header_fill = PatternFill("solid", fgColor="1a1a2e")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    thin = Side(style="thin", color="CCCCCC")
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    grade_fills = {
        "C": PatternFill("solid", fgColor="2a9d8f"),
        "D": PatternFill("solid", fgColor="e9c46a"),
        "E": PatternFill("solid", fgColor="e76f51"),
        "insufficient_data": PatternFill("solid", fgColor="aaaaaa"),
    }
    grade_fonts = {
        "C": Font(color="FFFFFF", bold=True, size=10),
        "D": Font(color="333333", bold=True, size=10),
        "E": Font(color="FFFFFF", bold=True, size=10),
        "insufficient_data": Font(color="444444", bold=True, size=10),
    }

    # Header row
    for col_idx, col in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = cell_border

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    # Data rows
    for row_idx, r in enumerate(rows, 2):
        grade = r.get("grade", "")
        for col_idx, col in enumerate(COLUMNS, 1):
            val = r.get(col, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = cell_border
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            if col == "grade" and grade in grade_fills:
                cell.fill = grade_fills[grade]
                cell.font = grade_fonts[grade]
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col == "score" and val is not None:
                cell.number_format = "0.0"
                if isinstance(val, (int, float)):
                    if val >= 60:
                        cell.font = Font(color="2a9d8f", bold=True, size=10)
                    elif val < 25:
                        cell.font = Font(color="e76f51", bold=True, size=10)

    # Column widths
    col_widths = {
        "product_name": 42, "score": 8, "grade": 14, "category": 18,
        "category_confidence": 12, "nova_proxy": 8, "confidence_score": 12,
        "evaluation_status": 14, "unresolved_flags": 14, "caps_applied": 32,
        "floors_applied": 26, "dominant_concern_family": 22,
        "dominant_positive_signal": 28, "structural_emptiness": 14,
        "product_id": 24,
    }
    for col_idx, col in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col, 14)

    ws.row_dimensions[1].height = 20

    wb.save(out_path)
    print(f"  XLSX -> {out_path}")


# ── HTML dashboard ───────────────────────────────────────────────────────────

def _trace_detail(t):
    """Return a dict with all the expanded trace info for one product."""
    ref = t.get("input_reference") or {}
    dim_scores = t.get("dimension_scores") or {}
    dim_weights = t.get("dimension_weights") or {}
    dim_notes = t.get("dimension_notes") or {}

    score_pipeline = []
    score_pipeline.append(f"Weighted dim score: {t.get('weighted_dimension_score')}")
    if t.get("caps_applied"):
        for c in t["caps_applied"]:
            score_pipeline.append(f"Cap applied: {c['rule']} → {c['cap']}")
    score_pipeline.append(f"Score after cap: {t.get('score_after_cap')}")
    if t.get("penalties_applied"):
        total = t.get("total_penalty_after_scaling")
        score_pipeline.append(f"Total penalty (scaled): -{total}")
    score_pipeline.append(f"Score after penalty: {t.get('score_after_penalty')}")
    if t.get("floors_applied"):
        for fl in t["floors_applied"]:
            score_pipeline.append(f"Floor applied: {fl.get('rule')} → {fl.get('floor_value')}")
    score_pipeline.append(f"Score after floors: {t.get('score_after_floors')}")
    if t.get("confidence_ceiling_applied"):
        score_pipeline.append(f"Confidence ceiling applied: {t['confidence_ceiling_applied']}")
    score_pipeline.append(f"FINAL: {t.get('final_score_estimate')} ({t.get('grade_estimate')})")

    dimensions = [
        {
            "name": k,
            "score": v,
            "weight": dim_weights.get(k),
            "note": dim_notes.get(k, ""),
        }
        for k, v in sorted(dim_scores.items(), key=lambda x: -(x[1] or 0))
    ]

    caps = t.get("caps_applied") or []
    penalties = t.get("penalties_applied") or []
    floors = t.get("floors_applied") or []
    conf_reductions = t.get("confidence_reductions") or []
    explanation = t.get("explanation_drivers") or []
    flags = t.get("unresolved_flags") or []

    return {
        "product_id": ref.get("canonical_product_id", ""),
        "score_pipeline": score_pipeline,
        "dimensions": dimensions,
        "caps": caps,
        "penalties": penalties,
        "floors": floors,
        "conf_reductions": conf_reductions,
        "explanation": explanation,
        "flags": flags,
        "category_instability": t.get("category_instability_flag"),
        "secondary_category": t.get("secondary_category"),
        "secondary_confidence": t.get("secondary_confidence"),
        "sugar_context_class": t.get("sugar_context_class"),
        "nova_evidence_for": t.get("nova_evidence_for") or [],
        "nova_evidence_against": t.get("nova_evidence_against") or [],
        "nova_uncertainty": t.get("nova_uncertainty_notes", ""),
        "scope_basis": t.get("scope_basis", ""),
        "context_note": t.get("context_note", ""),
    }


def export_html(rows, traces, out_path):
    # Build product table data and trace detail lookup
    trace_map = {}
    for t in traces:
        ref = t.get("input_reference") or {}
        pid = ref.get("canonical_product_id", "")
        trace_map[pid] = t

    table_data = []
    for r in rows:
        table_data.append({
            "product_id": r["product_id"],
            "product_name": r["product_name"],
            "score": r["score"],
            "grade": r["grade"],
            "category": r["category"],
            "category_confidence": r["category_confidence"],
            "nova_proxy": r["nova_proxy"],
            "confidence_score": r["confidence_score"],
            "evaluation_status": r["evaluation_status"],
            "unresolved_flags": r["unresolved_flags"],
            "caps_applied": r["caps_applied"],
            "floors_applied": r["floors_applied"],
            "dominant_concern_family": r["dominant_concern_family"],
            "dominant_positive_signal": r["dominant_positive_signal"],
            "structural_emptiness": r["structural_emptiness"],
        })

    detail_data = {}
    for r in rows:
        pid = r["product_id"]
        t = trace_map.get(pid, {})
        detail_data[pid] = _trace_detail(t)

    table_json = json.dumps(table_data, ensure_ascii=False)
    detail_json = json.dumps(detail_data, ensure_ascii=False)

    run_date = datetime.datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")

    html = f"""<!DOCTYPE html>
<html lang="he" dir="ltr">
<head>
<meta charset="UTF-8">
<title>BSIP2 Internal Review Dashboard</title>
<style>
  :root {{
    --bg: #f4f5f7;
    --surface: #ffffff;
    --header-bg: #1a1a2e;
    --header-fg: #e0e0e0;
    --border: #d0d0d0;
    --text: #222;
    --text-muted: #666;
    --accent: #4a6fa5;
    --grade-c: #2a9d8f;
    --grade-d: #e9c46a;
    --grade-e: #e76f51;
    --grade-i: #aaaaaa;
    --row-hover: #eef2f8;
    --expand-bg: #f8f9fc;
    --tag-bg: #e8edf5;
    --warn: #f28c28;
  }}
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: 'Segoe UI', Arial, sans-serif; background: var(--bg); color: var(--text); font-size: 13px; }}

  header {{
    background: var(--header-bg); color: var(--header-fg);
    padding: 12px 20px; display: flex; align-items: center; justify-content: space-between;
  }}
  header h1 {{ font-size: 16px; font-weight: 600; letter-spacing: 0.5px; }}
  header .meta {{ font-size: 11px; color: #888; }}

  .controls {{
    display: flex; flex-wrap: wrap; gap: 8px; padding: 12px 16px;
    background: var(--surface); border-bottom: 1px solid var(--border);
    position: sticky; top: 0; z-index: 100;
  }}
  .controls input, .controls select {{
    padding: 5px 10px; border: 1px solid var(--border); border-radius: 4px;
    font-size: 12px; background: var(--bg); color: var(--text);
  }}
  .controls input {{ width: 220px; }}
  .controls label {{ font-size: 11px; color: var(--text-muted); align-self: center; }}
  .count-badge {{
    margin-left: auto; font-size: 11px; color: var(--text-muted); align-self: center;
    background: var(--tag-bg); padding: 3px 8px; border-radius: 10px;
  }}

  .table-wrap {{ overflow-x: auto; padding: 0 0 40px 0; }}
  table {{ width: 100%; border-collapse: collapse; font-size: 12px; }}
  thead tr {{ background: var(--header-bg); color: var(--header-fg); }}
  thead th {{
    padding: 8px 10px; text-align: left; font-weight: 600; font-size: 11px;
    letter-spacing: 0.4px; cursor: pointer; white-space: nowrap;
    user-select: none; position: sticky; top: 53px; background: var(--header-bg);
  }}
  thead th:hover {{ background: #2a2a4a; }}
  thead th.sorted-asc::after {{ content: " ▲"; opacity: 0.7; }}
  thead th.sorted-desc::after {{ content: " ▼"; opacity: 0.7; }}

  tbody tr.data-row {{
    border-bottom: 1px solid var(--border); cursor: pointer;
    transition: background 0.1s;
  }}
  tbody tr.data-row:hover {{ background: var(--row-hover); }}
  tbody tr.data-row.expanded {{ background: #eef2f8; }}

  td {{ padding: 6px 10px; vertical-align: middle; }}
  td.name-cell {{ max-width: 260px; font-weight: 500; direction: rtl; text-align: right; }}
  td.score-cell {{ font-weight: 700; font-size: 13px; text-align: center; }}
  td.grade-cell {{ text-align: center; font-weight: 700; border-radius: 3px; padding: 4px 8px; }}
  td.num {{ text-align: center; }}
  td.flag-cell {{ color: var(--warn); font-weight: 600; text-align: center; }}
  td.muted {{ color: var(--text-muted); }}

  .grade-C {{ background: var(--grade-c); color: #fff; }}
  .grade-D {{ background: var(--grade-d); color: #333; }}
  .grade-E {{ background: var(--grade-e); color: #fff; }}
  .grade-insufficient_data {{ background: var(--grade-i); color: #444; }}
  .score-high {{ color: var(--grade-c); }}
  .score-low {{ color: var(--grade-e); }}

  /* Expand row */
  tr.expand-row {{ display: none; }}
  tr.expand-row.visible {{ display: table-row; }}
  tr.expand-row td {{
    padding: 0; background: var(--expand-bg);
    border-bottom: 2px solid var(--accent);
  }}
  .expand-inner {{ padding: 14px 18px; display: grid; grid-template-columns: 1fr 1fr 1fr; gap: 16px; }}

  .detail-section h4 {{
    font-size: 11px; text-transform: uppercase; letter-spacing: 0.5px;
    color: var(--text-muted); margin-bottom: 6px; border-bottom: 1px solid var(--border);
    padding-bottom: 3px;
  }}
  .detail-section ul {{ list-style: none; }}
  .detail-section li {{ padding: 2px 0; font-size: 11.5px; }}
  .detail-section li.warn {{ color: var(--warn); }}
  .detail-section li.good {{ color: var(--grade-c); }}

  .dim-bar-wrap {{ display: flex; align-items: center; gap: 6px; margin-bottom: 3px; }}
  .dim-label {{ width: 140px; font-size: 11px; color: var(--text-muted); }}
  .dim-bar-bg {{ flex: 1; background: #e0e0e0; border-radius: 2px; height: 8px; }}
  .dim-bar {{ height: 8px; border-radius: 2px; background: var(--accent); }}
  .dim-val {{ width: 32px; font-size: 11px; text-align: right; font-weight: 600; }}

  .pipeline-list {{ font-family: monospace; font-size: 11px; }}
  .pipeline-list li {{ padding: 1px 0; }}
  .pipeline-list li:last-child {{ font-weight: 700; color: var(--accent); }}

  .tag {{
    display: inline-block; background: var(--tag-bg); border-radius: 3px;
    padding: 1px 6px; font-size: 10px; margin: 1px; color: var(--text-muted);
  }}
  .tag.warn {{ background: #fff0e0; color: var(--warn); }}
  .tag.good {{ background: #e0f5f0; color: #2a9d8f; }}

  .expand-toggle {{ color: var(--accent); font-size: 11px; padding: 2px 6px; }}

  /* NOVA badge */
  .nova {{ display: inline-block; width: 22px; height: 22px; border-radius: 50%;
    text-align: center; line-height: 22px; font-weight: 700; font-size: 11px; color: #fff; }}
  .nova-2 {{ background: #2a9d8f; }}
  .nova-3 {{ background: #e9c46a; color: #333; }}
  .nova-4 {{ background: #e76f51; }}

  .no-results {{ text-align: center; padding: 40px; color: var(--text-muted); font-size: 14px; }}
</style>
</head>
<body>
<header>
  <div>
    <h1>BSIP2 Internal Review Dashboard &nbsp; <span style="font-weight:300;font-size:12px">proto_v0</span></h1>
    <div class="meta">Run: {run_date} &nbsp;|&nbsp; bsip2_concept_v1 + SRC-v1 &nbsp;|&nbsp; Signal hygiene pass</div>
  </div>
  <div class="meta" style="text-align:right">Internal diagnostics — not for distribution</div>
</header>

<div class="controls">
  <label>Search:</label>
  <input type="text" id="searchBox" placeholder="Product name, ID, category…" oninput="applyFilters()">
  <label>Grade:</label>
  <select id="gradeFilter" onchange="applyFilters()">
    <option value="">All</option>
    <option value="C">C</option>
    <option value="D">D</option>
    <option value="E">E</option>
    <option value="insufficient_data">Insufficient data</option>
  </select>
  <label>Category:</label>
  <select id="catFilter" onchange="applyFilters()">
    <option value="">All</option>
    <option value="snack_bar_granola">snack_bar_granola</option>
    <option value="whole_food_fat">whole_food_fat</option>
    <option value="cereal">cereal</option>
    <option value="dairy_protein">dairy_protein</option>
  </select>
  <label>NOVA:</label>
  <select id="novaFilter" onchange="applyFilters()">
    <option value="">All</option>
    <option value="2">2</option>
    <option value="3">3</option>
    <option value="4">4</option>
  </select>
  <label>Flags:</label>
  <select id="flagFilter" onchange="applyFilters()">
    <option value="">All</option>
    <option value="has_flags">Has unresolved flags</option>
    <option value="no_flags">No flags</option>
  </select>
  <span class="count-badge" id="countBadge">— products</span>
</div>

<div class="table-wrap">
<table id="mainTable">
<thead>
  <tr>
    <th style="width:22px"></th>
    <th data-col="product_name">Product Name</th>
    <th data-col="score" class="sorted-desc">Score</th>
    <th data-col="grade">Grade</th>
    <th data-col="category">Category</th>
    <th data-col="category_confidence">Cat Conf</th>
    <th data-col="nova_proxy">NOVA</th>
    <th data-col="confidence_score">Confidence</th>
    <th data-col="evaluation_status">Eval Status</th>
    <th data-col="unresolved_flags">Flags</th>
    <th data-col="dominant_concern_family">Concern Family</th>
    <th data-col="dominant_positive_signal">Positive Signal</th>
    <th data-col="structural_emptiness">Struct Empty</th>
    <th data-col="product_id">Product ID</th>
  </tr>
</thead>
<tbody id="tableBody"></tbody>
</table>
<div class="no-results" id="noResults" style="display:none">No products match the current filters.</div>
</div>

<script>
const TABLE_DATA = {table_json};
const DETAIL_DATA = {detail_json};

let sortCol = "score";
let sortDir = -1; // -1=desc, 1=asc
let filteredData = [...TABLE_DATA];

function scoreColor(score) {{
  if (score === null || score === undefined) return "";
  if (score >= 55) return "score-high";
  if (score < 25) return "score-low";
  return "";
}}

function novaHtml(n) {{
  if (!n) return "—";
  return `<span class="nova nova-${{n}}">${{n}}</span>`;
}}

function gradeHtml(g) {{
  if (!g) return "—";
  return `<span class="grade-cell grade-${{g}}">${{g === "insufficient_data" ? "n/d" : g}}</span>`;
}}

function flagHtml(n) {{
  if (!n) return '<span class="muted">0</span>';
  return `<span class="flag-cell">⚑ ${{n}}</span>`;
}}

function fmt(v) {{
  if (v === null || v === undefined || v === "") return "—";
  if (typeof v === "number") return v.toFixed ? v.toFixed(1) : v;
  return v;
}}

function renderTable() {{
  const tbody = document.getElementById("tableBody");
  tbody.innerHTML = "";
  if (filteredData.length === 0) {{
    document.getElementById("noResults").style.display = "";
    document.getElementById("countBadge").textContent = "0 products";
    return;
  }}
  document.getElementById("noResults").style.display = "none";
  document.getElementById("countBadge").textContent = filteredData.length + " products";

  filteredData.forEach((row, i) => {{
    const pid = row.product_id;
    const dataRow = document.createElement("tr");
    dataRow.className = "data-row";
    dataRow.dataset.pid = pid;
    dataRow.onclick = () => toggleExpand(pid, dataRow);

    const scoreClass = scoreColor(row.score);

    dataRow.innerHTML = `
      <td class="expand-toggle">▶</td>
      <td class="name-cell" title="${{html_escape(row.product_id)}}">${{html_escape(row.product_name)}}</td>
      <td class="score-cell ${{scoreClass}}">${{row.score !== null && row.score !== undefined ? Number(row.score).toFixed(1) : "—"}}</td>
      <td class="num">${{gradeHtml(row.grade)}}</td>
      <td class="muted">${{html_escape(row.category)}}</td>
      <td class="num">${{row.category_confidence !== null && row.category_confidence !== undefined ? Number(row.category_confidence).toFixed(2) : "—"}}</td>
      <td class="num">${{novaHtml(row.nova_proxy)}}</td>
      <td class="num">${{row.confidence_score !== null && row.confidence_score !== undefined ? Number(row.confidence_score).toFixed(0) : "—"}}</td>
      <td class="muted">${{html_escape(row.evaluation_status)}}</td>
      <td class="num">${{flagHtml(row.unresolved_flags)}}</td>
      <td class="muted">${{html_escape(row.dominant_concern_family || "—")}}</td>
      <td class="muted" style="font-size:10px">${{html_escape(row.dominant_positive_signal || "—")}}</td>
      <td class="num ${{row.structural_emptiness === "yes" ? "flag-cell" : "muted"}}">${{html_escape(row.structural_emptiness)}}</td>
      <td class="muted" style="font-size:10px">${{html_escape(row.product_id)}}</td>
    `;
    tbody.appendChild(dataRow);

    // Expand row (hidden by default)
    const expandRow = document.createElement("tr");
    expandRow.className = "expand-row";
    expandRow.id = "expand-" + pid;
    const expandTd = document.createElement("td");
    expandTd.colSpan = 14;
    expandTd.innerHTML = buildExpandHtml(pid);
    expandRow.appendChild(expandTd);
    tbody.appendChild(expandRow);
  }});
}}

function html_escape(s) {{
  if (s === null || s === undefined) return "—";
  return String(s)
    .replace(/&/g, "&amp;")
    .replace(/</g, "&lt;")
    .replace(/>/g, "&gt;")
    .replace(/"/g, "&quot;");
}}

function buildExpandHtml(pid) {{
  const d = DETAIL_DATA[pid];
  if (!d) return "<div style='padding:10px;color:#999'>No detail available.</div>";

  // Score pipeline
  const pipeline = (d.score_pipeline || []).map(s => `<li>${{html_escape(s)}}</li>`).join("");

  // Dimensions
  const dims = (d.dimensions || []).map(dim => {{
    const pct = Math.min(100, Math.round((dim.score || 0)));
    const barColor = dim.score >= 70 ? "#2a9d8f" : dim.score >= 45 ? "#e9c46a" : "#e76f51";
    return `<div class="dim-bar-wrap">
      <span class="dim-label">${{html_escape(dim.name)}}</span>
      <div class="dim-bar-bg"><div class="dim-bar" style="width:${{pct}}%;background:${{barColor}}"></div></div>
      <span class="dim-val">${{dim.score !== null ? Number(dim.score).toFixed(1) : "—"}}</span>
    </div>`;
  }}).join("");

  // Caps
  const caps = (d.caps || []).length
    ? (d.caps || []).map(c => `<li class="warn">${{html_escape(c.rule)}} → cap ${{c.cap}}</li>`).join("")
    : "<li class='muted'>None</li>";

  // Penalties
  const pens = (d.penalties || []).length
    ? (d.penalties || []).map(p => `<li class="warn">${{html_escape(p.rule)}} −${{p.amount}}${{p.note ? ' (' + html_escape(p.note) + ')' : ''}}</li>`).join("")
    : "<li class='muted'>None</li>";

  // Floors
  const floors = (d.floors || []).length
    ? (d.floors || []).map(f => `<li class="good">${{html_escape(f.rule || f.floor_type || "?")}} → ${{f.floor_value}}</li>`).join("")
    : "<li class='muted'>None</li>";

  // Confidence reductions
  const confs = (d.conf_reductions || []).length
    ? (d.conf_reductions || []).map(c => `<li>${{html_escape(c.factor)}}: ${{c.reduction}}</li>`).join("")
    : "<li class='muted'>None</li>";

  // Flags
  const flags = (d.flags || []).length
    ? (d.flags || []).map(f => `<li class="warn">⚑ ${{html_escape(f)}}</li>`).join("")
    : "<li class='muted'>No unresolved flags</li>";

  // Explanation drivers
  const expl = (d.explanation || []).map(e => `<li>${{html_escape(e)}}</li>`).join("");

  // NOVA evidence
  const novaFor = (d.nova_evidence_for || []).map(e => `<span class="tag good">${{html_escape(e)}}</span>`).join(" ");
  const novaAgainst = (d.nova_evidence_against || []).map(e => `<span class="tag warn">${{html_escape(e)}}</span>`).join(" ");

  // Category instability
  let catInstability = '<span class="muted">None</span>';
  if (d.category_instability) {{
    catInstability = `<span class="tag warn">Instability: primary vs secondary=${{html_escape(d.secondary_category)}} (${{d.secondary_confidence !== null ? Number(d.secondary_confidence).toFixed(2) : "—"}})</span>`;
  }}

  return `<div class="expand-inner">
    <div class="detail-section">
      <h4>Score Pipeline</h4>
      <ul class="pipeline-list">${{pipeline}}</ul>
      <br>
      <h4>Explanation Drivers</h4>
      <ul>${{expl || "<li class='muted'>—</li>"}}</ul>
      <br>
      <h4>Unresolved Flags</h4>
      <ul>${{flags}}</ul>
    </div>
    <div class="detail-section">
      <h4>Dimension Scores</h4>
      ${{dims}}
      <br>
      <h4>Caps Applied</h4>
      <ul>${{caps}}</ul>
      <br>
      <h4>Penalties Applied</h4>
      <ul>${{pens}}</ul>
      <br>
      <h4>Floors Applied</h4>
      <ul>${{floors}}</ul>
    </div>
    <div class="detail-section">
      <h4>Confidence Reductions</h4>
      <ul>${{confs}}</ul>
      <br>
      <h4>NOVA Evidence</h4>
      <div style="margin-bottom:4px">For: ${{novaFor || '<span class="muted">—</span>'}}</div>
      <div>Against: ${{novaAgainst || '<span class="muted">—</span>'}}</div>
      ${{d.nova_uncertainty ? `<div style="margin-top:4px;font-size:11px;color:#888">${{html_escape(d.nova_uncertainty)}}</div>` : ''}}
      <br>
      <h4>Category Instability</h4>
      ${{catInstability}}
      <br>
      <h4>Context</h4>
      <ul>
        ${{d.sugar_context_class ? `<li>Sugar context: ${{html_escape(d.sugar_context_class)}}</li>` : ''}}
        ${{d.scope_basis ? `<li>Scope: ${{html_escape(d.scope_basis)}}</li>` : ''}}
        ${{d.context_note ? `<li class="warn">Note: ${{html_escape(d.context_note)}}</li>` : ''}}
      </ul>
    </div>
  </div>`;
}}

function toggleExpand(pid, dataRow) {{
  const expandRow = document.getElementById("expand-" + pid);
  if (!expandRow) return;
  const isOpen = expandRow.classList.contains("visible");
  // Close all others
  document.querySelectorAll(".expand-row.visible").forEach(r => {{
    r.classList.remove("visible");
    const dr = r.previousElementSibling;
    if (dr) {{ dr.classList.remove("expanded"); dr.querySelector(".expand-toggle").textContent = "▶"; }}
  }});
  if (!isOpen) {{
    expandRow.classList.add("visible");
    dataRow.classList.add("expanded");
    dataRow.querySelector(".expand-toggle").textContent = "▼";
  }}
}}

// Sorting
document.querySelectorAll("thead th[data-col]").forEach(th => {{
  th.addEventListener("click", () => {{
    const col = th.dataset.col;
    if (sortCol === col) sortDir = -sortDir;
    else {{ sortCol = col; sortDir = -1; }}
    document.querySelectorAll("thead th").forEach(t => t.classList.remove("sorted-asc","sorted-desc"));
    th.classList.add(sortDir === -1 ? "sorted-desc" : "sorted-asc");
    sortData();
    renderTable();
  }});
}});

function sortData() {{
  filteredData.sort((a, b) => {{
    let va = a[sortCol], vb = b[sortCol];
    if (va === null || va === undefined) va = sortDir === -1 ? -Infinity : Infinity;
    if (vb === null || vb === undefined) vb = sortDir === -1 ? -Infinity : Infinity;
    if (typeof va === "string") va = va.toLowerCase();
    if (typeof vb === "string") vb = vb.toLowerCase();
    if (va < vb) return -sortDir;
    if (va > vb) return sortDir;
    return 0;
  }});
}}

function applyFilters() {{
  const search = document.getElementById("searchBox").value.toLowerCase();
  const grade = document.getElementById("gradeFilter").value;
  const cat = document.getElementById("catFilter").value;
  const nova = document.getElementById("novaFilter").value;
  const flag = document.getElementById("flagFilter").value;

  filteredData = TABLE_DATA.filter(row => {{
    if (grade && row.grade !== grade) return false;
    if (cat && row.category !== cat) return false;
    if (nova && String(row.nova_proxy) !== nova) return false;
    if (flag === "has_flags" && !row.unresolved_flags) return false;
    if (flag === "no_flags" && row.unresolved_flags > 0) return false;
    if (search) {{
      const hay = (row.product_name + " " + row.product_id + " " + row.category).toLowerCase();
      if (!hay.includes(search)) return false;
    }}
    return true;
  }});
  sortData();
  renderTable();
}}

// Initial render
sortData();
renderTable();
</script>
</body>
</html>
"""

    with open(out_path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"  HTML -> {out_path}")


# ── main ─────────────────────────────────────────────────────────────────────

def main():
    REVIEW_ROOT.mkdir(parents=True, exist_ok=True)
    print(f"Loading traces from {TRACE_ROOT} …")
    traces = load_traces()
    print(f"  Loaded {len(traces)} traces.")

    rows = build_all_rows(traces)

    csv_path = REVIEW_ROOT / "bsip2_master_products.csv"
    xlsx_path = REVIEW_ROOT / "bsip2_master_products.xlsx"
    html_path = REVIEW_ROOT / "review_dashboard.html"

    export_csv(rows, csv_path)
    export_xlsx(rows, xlsx_path)
    export_html(rows, traces, html_path)

    print(f"\nDone. {len(rows)} products written.")
    print(f"  Sufficient data:    {sum(1 for r in rows if r['grade'] != 'insufficient_data')}")
    print(f"  Insufficient data:  {sum(1 for r in rows if r['grade'] == 'insufficient_data')}")
    print(f"  With flags:         {sum(1 for r in rows if r['unresolved_flags'] > 0)}")


if __name__ == "__main__":
    main()
