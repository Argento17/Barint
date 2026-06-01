from pathlib import Path
import json
from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent

OUTPUT_DIR = BASE_DIR / "outputs"
GROUND_TRUTH_FILE = BASE_DIR / "ground_truth.xlsx"

EXCLUDED_PRODUCTS = {"product_010"}

FIELDS = [
    "energy_kcal_100g",
    "fat_g_100g",
    "saturated_fat_g_100g",
    "carbohydrates_g_100g",
    "fiber_g_100g",
    "protein_g_100g",
    "sodium_mg_100g",
    "cholesterol_mg_100g",
    "sugar_tbsp_100g"
]


def normalize_number(value):
    if value is None or value == "":
        return None

    try:
        return float(value)
    except Exception:
        return None


def load_ground_truth():
    wb = load_workbook(GROUND_TRUTH_FILE)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]

    data = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(headers, row))
        product_id = row_dict.get("product_id")

        if not product_id:
            continue

        if product_id in EXCLUDED_PRODUCTS:
            continue

        data[product_id] = {
            field: normalize_number(row_dict.get(field))
            for field in FIELDS
        }

    return data


def load_prediction(product_id):
    path = OUTPUT_DIR / f"{product_id}.json"

    if not path.exists():
        return None

    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)

    return data.get("extracted", {}).get("nutrition_per_100g", {})


def compare(expected, predicted):
    if expected is None and predicted is None:
        return "both_missing"

    if expected is None and predicted is not None:
        return "unexpected_value"

    if expected is not None and predicted is None:
        return "missing_prediction"

    tolerance = max(0.2, expected * 0.03)

    if abs(expected - predicted) <= tolerance:
        return "match"

    return "mismatch"


def main():
    ground_truth = load_ground_truth()

    total = 0
    matches = 0

    print("\nParser evaluation\n")

    for product_id, expected_values in ground_truth.items():
        predicted_values = load_prediction(product_id)

        if predicted_values is None:
            print(f"{product_id}: missing output JSON")
            continue

        print(f"\n--- {product_id} ---")

        for field in FIELDS:
            expected = expected_values.get(field)
            predicted = normalize_number(predicted_values.get(field))
            status = compare(expected, predicted)

            if expected is not None:
                total += 1

                if status == "match":
                    matches += 1

            print(
                f"{field}: expected={expected} | predicted={predicted} | {status}"
            )

    accuracy = (matches / total * 100) if total else 0

    print("\n====================")
    print(f"Matched fields: {matches}/{total}")
    print(f"Accuracy: {accuracy:.1f}%")
    print("====================\n")


if __name__ == "__main__":
    main()