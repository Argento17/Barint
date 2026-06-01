#!/usr/bin/env python3
"""
BSIP1 batch canonicalization test — run 001.

Reads all product.json files from Carrefour and Yohananof BSIP0 output folders,
groups by barcode, canonicalizes each group, and writes the two-file output
(compact product + audit sidecar) to batch_test_001/output/.

Also generates six analytical reports in batch_test_001/reports/:
  1. batch_report.md
  2. barcode_quality_report.xlsx
  3. nutrition_consistency_warnings.xlsx
  4. ingredient_quality_warnings.xlsx
  5. conflict_breakdown.xlsx
  6. fuzzy_candidate_queue.json

Run:
    cd C:\\Bari
    .venv\\Scripts\\python 03_operations\\bsip1\\run_001\\batch_test_001.py
"""

from __future__ import annotations

import json
import logging
import sys
from pathlib import Path

BASE_DIR = Path(__file__).resolve().parent
BSIP1_ROOT = BASE_DIR.parent
sys.path.insert(0, str(BSIP1_ROOT))

from core.reader import read_bsip0_root, group_by_barcode
from core.merger import canonicalize

OUTPUT_DIR   = BASE_DIR / 'output'
REPORTS_DIR  = BASE_DIR / 'reports'
BSIP0_ROOT   = Path(r'C:\Bari\03_operations\bsip0\scrape')
RETAILER_IDS = ['carrefour', 'yohananof']
SCHEMAS_DIR  = BSIP1_ROOT / 'schemas'

OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
REPORTS_DIR.mkdir(parents=True, exist_ok=True)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s %(levelname)-8s %(message)s',
    handlers=[
        logging.FileHandler(OUTPUT_DIR / 'batch_test_001.log', encoding='utf-8'),
        logging.StreamHandler(),
    ],
)
log = logging.getLogger('bsip1_batch')


def _validate(doc: dict, schema_filename: str, label: str) -> str:
    path = SCHEMAS_DIR / schema_filename
    if not path.exists():
        return 'schema_file_not_found'
    try:
        import jsonschema
        schema = json.loads(path.read_text(encoding='utf-8'))
        errors = list(jsonschema.Draft202012Validator(schema).iter_errors(doc))
        if errors:
            for e in errors:
                log.error('%s schema error [%s]: %s', label, list(e.path), e.message)
            return f'INVALID ({len(errors)} errors)'
        return 'VALID'
    except ImportError:
        return 'jsonschema_not_installed'


def _write_pair(product: dict, audit: dict) -> tuple[Path, Path]:
    pid        = product['canonical_product_id']
    prod_path  = OUTPUT_DIR / f'{pid}.json'
    audit_path = OUTPUT_DIR / product['audit_ref']
    prod_path.write_text(json.dumps(product, ensure_ascii=False, indent=2), encoding='utf-8')
    audit_path.write_text(json.dumps(audit,   ensure_ascii=False, indent=2), encoding='utf-8')
    return prod_path, audit_path


# ── report helpers ────────────────────────────────────────────────────────────

_HEADER_FILL = None  # populated only when openpyxl available


def _make_wb_ws(title: str):
    """Create a new workbook and worksheet with a styled header row."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title
    return wb, ws


def _style_header(ws, cols: list[str]) -> None:
    """Write and style the header row."""
    fill = PatternFill('solid', fgColor='1F4E79')
    font = Font(bold=True, color='FFFFFF')
    for col_idx, col_name in enumerate(cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(wrap_text=True)
    ws.row_dimensions[1].height = 20


def _save_or_csv(wb, ws, rows: list[list], cols: list[str], xlsx_path: Path) -> None:
    """Write rows to xlsx if openpyxl available, else write csv fallback."""
    if HAS_OPENPYXL:
        _style_header(ws, cols)
        for row in rows:
            ws.append(row)
        for col_cells in ws.columns:
            max_len = max(len(str(c.value or '')) for c in col_cells)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 60)
        wb.save(xlsx_path)
    else:
        import csv
        csv_path = xlsx_path.with_suffix('.csv')
        with open(csv_path, 'w', newline='', encoding='utf-8') as f:
            w = csv.writer(f)
            w.writerow(cols)
            w.writerows(rows)
        log.warning('openpyxl not installed — wrote CSV fallback: %s', csv_path)


# ── report generators ─────────────────────────────────────────────────────────

def write_barcode_quality_report(products: list[dict]) -> None:
    cols = [
        'canonical_product_id', 'barcode', 'barcode_length',
        'barcode_validation_status', 'barcode_confidence_reason',
        'source_retailers', 'observation_count',
    ]
    rows = []
    for p in products:
        bc = p.get('barcode') or ''
        rows.append([
            p['canonical_product_id'],
            bc,
            len(bc) if bc else '',
            p.get('barcode_validation_status', ''),
            p.get('barcode_confidence_reason', ''),
            ', '.join(p.get('source_retailers', [])),
            p['confidence']['observation_count'],
        ])
    wb, ws = _make_wb_ws('Barcode Quality') if HAS_OPENPYXL else (None, None)
    _save_or_csv(wb, ws, rows, cols, REPORTS_DIR / 'barcode_quality_report.xlsx')


def write_nutrition_warnings_report(products: list[dict]) -> None:
    cols = [
        'canonical_product_id', 'barcode',
        'nutrition_consistency_status', 'nutrition_basis_claimed',
        'nutrition_basis_detected', 'warning_index', 'warning_text',
    ]
    rows = []
    for p in products:
        warnings = p.get('nutrition_consistency_warnings') or []
        status   = p.get('nutrition_consistency_status', '')
        if warnings:
            for idx, w in enumerate(warnings, 1):
                rows.append([
                    p['canonical_product_id'],
                    p.get('barcode', ''),
                    status,
                    p.get('nutrition_basis_claimed', ''),
                    p.get('nutrition_basis_detected', ''),
                    idx,
                    w,
                ])
        else:
            rows.append([
                p['canonical_product_id'],
                p.get('barcode', ''),
                status,
                p.get('nutrition_basis_claimed', ''),
                p.get('nutrition_basis_detected', ''),
                '',
                '(no warnings)',
            ])
    wb, ws = _make_wb_ws('Nutrition Warnings') if HAS_OPENPYXL else (None, None)
    _save_or_csv(wb, ws, rows, cols, REPORTS_DIR / 'nutrition_consistency_warnings.xlsx')


def write_ingredient_warnings_report(products: list[dict]) -> None:
    cols = [
        'canonical_product_id', 'barcode',
        'ingredient_text_quality', 'warning_index', 'warning_text',
        'ingredients_preview',
    ]
    rows = []
    for p in products:
        warnings = p.get('ingredient_warnings') or []
        quality  = p.get('ingredient_text_quality', '')
        preview  = (p.get('ingredients_text_he') or '')[:80]
        if warnings:
            for idx, w in enumerate(warnings, 1):
                rows.append([
                    p['canonical_product_id'],
                    p.get('barcode', ''),
                    quality,
                    idx,
                    w,
                    preview,
                ])
        else:
            rows.append([
                p['canonical_product_id'],
                p.get('barcode', ''),
                quality,
                '',
                '(no warnings)',
                preview,
            ])
    wb, ws = _make_wb_ws('Ingredient Warnings') if HAS_OPENPYXL else (None, None)
    _save_or_csv(wb, ws, rows, cols, REPORTS_DIR / 'ingredient_quality_warnings.xlsx')


def write_conflict_breakdown_report(products: list[dict]) -> None:
    cols = [
        'canonical_product_id', 'barcode',
        'total_conflicts', 'has_unresolved',
        'identity_count', 'nutrition_count',
        'ingredient_count', 'labeling_count', 'completeness_count',
        'identity_fields', 'nutrition_fields',
    ]
    rows = []
    for p in products:
        cs = p.get('conflicts_summary', {})
        rows.append([
            p['canonical_product_id'],
            p.get('barcode', ''),
            cs.get('count', 0),
            cs.get('has_unresolved', False),
            len(cs.get('identity_conflicts', [])),
            len(cs.get('nutrition_conflicts', [])),
            len(cs.get('ingredient_conflicts', [])),
            len(cs.get('labeling_conflicts', [])),
            len(cs.get('completeness_conflicts', [])),
            ', '.join(cs.get('identity_conflicts', [])),
            ', '.join(cs.get('nutrition_conflicts', [])),
        ])
    wb, ws = _make_wb_ws('Conflict Breakdown') if HAS_OPENPYXL else (None, None)
    _save_or_csv(wb, ws, rows, cols, REPORTS_DIR / 'conflict_breakdown.xlsx')


def write_fuzzy_candidate_queue(no_barcode: list[dict]) -> None:
    candidates = []
    for o in no_barcode:
        pkg  = o.get('package_size_parsed', {})
        ing  = (o.get('ingredients_raw_he') or '')
        candidates.append({
            'queue_reason':             'no_barcode_in_source',
            'retailer_id':              o.get('retailer_id', ''),
            'folder_path':              o.get('folder_path', ''),
            'product_name_raw':         o.get('product_name_raw', ''),
            'brand':                    o.get('brand'),
            'package_size_raw':         o.get('package_size_raw', ''),
            'package_size_g':           pkg.get('package_size_g'),
            'unit_count':               pkg.get('unit_count'),
            'unit_size_g':              pkg.get('unit_size_g'),
            'barcode_source':           o.get('barcode_source', ''),
            'scraped_at':               o.get('scraped_at', ''),
            'ingredients_raw_he_preview': ing[:100] + ('...' if len(ing) > 100 else ''),
            'match_hints': {
                'name_tokens': o.get('product_name_raw', '').split()[:6],
                'brand_lower': (o.get('brand') or '').lower().strip(),
            },
        })

    queue_path = REPORTS_DIR / 'fuzzy_candidate_queue.json'
    queue_path.write_text(
        json.dumps(
            {
                'generated_by': 'bsip1_batch_test_001',
                'total_candidates': len(candidates),
                'note': 'No-barcode observations queued for manual fuzzy matching. Do NOT auto-merge.',
                'candidates': candidates,
            },
            ensure_ascii=False, indent=2,
        ),
        encoding='utf-8',
    )
    log.info('fuzzy_candidate_queue.json: %d candidates → %s', len(candidates), queue_path)


def write_observation_quality_report(obs_rows: list[dict]) -> None:
    cols = [
        'canonical_product_id', 'barcode', 'retailer_id',
        'observation_quality_score', 'observation_quality_level',
        'scrape_mode', 'barcode_source', 'nutrition_completeness',
        'quality_signals',
    ]
    rows = [
        [
            r['canonical_product_id'],
            r.get('barcode', ''),
            r['retailer_id'],
            r['observation_quality_score'],
            r['observation_quality_level'],
            r['scrape_mode'],
            r.get('barcode_source', ''),
            r.get('nutrition_completeness', ''),
            ', '.join(r.get('observation_quality_signals', [])),
        ]
        for r in obs_rows
    ]
    wb, ws = _make_wb_ws('Observation Quality') if HAS_OPENPYXL else (None, None)
    _save_or_csv(wb, ws, rows, cols, REPORTS_DIR / 'observation_quality_report.xlsx')


def write_canonical_trust_report(products: list[dict]) -> None:
    cols = [
        'canonical_product_id', 'barcode',
        'canonical_trust_score', 'canonical_trust_level',
        'canonical_risk_flags',
        'identity_confidence', 'barcode_validation_status',
        'nutrition_consistency_status', 'ingredient_text_quality',
        'observation_count',
    ]
    rows = [
        [
            p['canonical_product_id'],
            p.get('barcode', ''),
            p.get('canonical_trust_score', ''),
            p.get('canonical_trust_level', ''),
            ', '.join(p.get('canonical_risk_flags', [])),
            p['confidence']['identity_confidence'],
            p.get('barcode_validation_status', ''),
            p.get('nutrition_consistency_status', ''),
            p.get('ingredient_text_quality', ''),
            p['confidence']['observation_count'],
        ]
        for p in products
    ]
    wb, ws = _make_wb_ws('Canonical Trust') if HAS_OPENPYXL else (None, None)
    _save_or_csv(wb, ws, rows, cols, REPORTS_DIR / 'canonical_trust_report.xlsx')


def write_high_risk_products_report(products: list[dict]) -> None:
    # All products with trust_level == 'low' OR at least one of the severe risk flags
    _severe = frozenset({
        'low_identity_confidence', 'nutrition_suspicious', 'unresolved_conflicts',
        'corrupted_ingredient_text', 'absent_or_weak_barcode', 'high_conflict_count',
    })
    risky = [
        p for p in products
        if p.get('canonical_trust_level') == 'low'
        or bool(_severe & set(p.get('canonical_risk_flags', [])))
    ]

    cols = [
        'canonical_product_id', 'barcode',
        'canonical_trust_score', 'canonical_trust_level',
        'canonical_risk_flags',
        'nutrition_consistency_status', 'ingredient_text_quality',
        'barcode_validation_status', 'observation_count',
        'total_conflicts',
    ]
    rows = [
        [
            p['canonical_product_id'],
            p.get('barcode', ''),
            p.get('canonical_trust_score', ''),
            p.get('canonical_trust_level', ''),
            ', '.join(p.get('canonical_risk_flags', [])),
            p.get('nutrition_consistency_status', ''),
            p.get('ingredient_text_quality', ''),
            p.get('barcode_validation_status', ''),
            p['confidence']['observation_count'],
            p['conflicts_summary']['count'],
        ]
        for p in risky
    ]
    wb, ws = _make_wb_ws('High Risk Products') if HAS_OPENPYXL else (None, None)
    _save_or_csv(wb, ws, rows, cols, REPORTS_DIR / 'high_risk_products.xlsx')
    log.info('high_risk_products.xlsx: %d products', len(risky))


def write_trust_distribution_md(
    products: list[dict],
    obs_rows: list[dict],
    run_date: str,
    obs_count_stats: dict,
) -> None:
    # Trust level distribution
    trust_dist: dict[str, list[dict]] = {'high': [], 'medium': [], 'low': []}
    for p in products:
        lvl = p.get('canonical_trust_level', 'low')
        trust_dist.setdefault(lvl, []).append(p)

    # Risk flag frequency
    flag_freq: dict[str, int] = {}
    for p in products:
        for flag in p.get('canonical_risk_flags', []):
            flag_freq[flag] = flag_freq.get(flag, 0) + 1

    # Observation signal frequency
    sig_freq: dict[str, int] = {}
    for r in obs_rows:
        for sig in r.get('observation_quality_signals', []):
            sig_freq[sig] = sig_freq.get(sig, 0) + 1

    # Average trust score
    scores = [p.get('canonical_trust_score') or 0.0 for p in products]
    avg_score = round(sum(scores) / len(scores), 3) if scores else 0.0
    min_score = round(min(scores), 3) if scores else 0.0
    max_score = round(max(scores), 3) if scores else 0.0

    lines = [
        '# BSIP1 Trust Distribution Summary',
        '',
        f'**Run date:** {run_date}  ',
        f'**Products evaluated:** {len(products)}  ',
        f'**Observations — total:** {obs_count_stats["total"]}  ',
        f'**Observations — trust-scored:** {obs_count_stats["scored"]}  ',
        f'**Observations — excluded from scoring:** {obs_count_stats["excluded"]}  ',
        f'**Exclusion reason:** {obs_count_stats["exclusion_reason"]}',
        '',
        '---',
        '',
        '## Canonical Trust Distribution',
        '',
        f'| Trust Level | Count | % |',
        f'|-------------|-------|---|',
    ]
    for lvl in ('high', 'medium', 'low'):
        cnt = len(trust_dist.get(lvl, []))
        pct = round(cnt / len(products) * 100, 1) if products else 0
        lines.append(f'| **{lvl}** | {cnt} | {pct}% |')

    lines += [
        '',
        f'Average canonical trust score: **{avg_score}**  ',
        f'Range: {min_score} – {max_score}',
        '',
        '---',
        '',
        '## High Trust Products',
        '',
        f'{len(trust_dist.get("high", []))} products with canonical_trust_level = high',
        '',
    ]
    for p in sorted(trust_dist.get('high', []), key=lambda x: -(x.get('canonical_trust_score') or 0)):
        flags = p.get('canonical_risk_flags', [])
        flag_str = ' (' + ', '.join(flags) + ')' if flags else ''
        lines.append(
            f'- `{p["canonical_product_id"]}` — score {p.get("canonical_trust_score", "?")}'
            f'{flag_str}'
        )

    lines += [
        '',
        '---',
        '',
        '## Medium Trust Products',
        '',
        f'{len(trust_dist.get("medium", []))} products with canonical_trust_level = medium',
        '',
    ]
    for p in sorted(trust_dist.get('medium', []), key=lambda x: -(x.get('canonical_trust_score') or 0)):
        flags = p.get('canonical_risk_flags', [])
        flag_str = ' (' + ', '.join(flags[:3]) + (', …' if len(flags) > 3 else '') + ')' if flags else ''
        lines.append(
            f'- `{p["canonical_product_id"]}` — score {p.get("canonical_trust_score", "?")}'
            f'{flag_str}'
        )

    lines += [
        '',
        '---',
        '',
        '## Low Trust Products',
        '',
        f'{len(trust_dist.get("low", []))} products with canonical_trust_level = low',
        '',
    ]
    for p in sorted(trust_dist.get('low', []), key=lambda x: (x.get('canonical_trust_score') or 0)):
        flags = p.get('canonical_risk_flags', [])
        flag_str = ' (' + ', '.join(flags[:3]) + (', …' if len(flags) > 3 else '') + ')' if flags else ''
        lines.append(
            f'- `{p["canonical_product_id"]}` — score {p.get("canonical_trust_score", "?")}'
            f'{flag_str}'
        )

    lines += [
        '',
        '---',
        '',
        '## Top Causes of Trust Degradation',
        '',
        'Canonical risk flag frequency across all products:',
        '',
        '| Risk Flag | Products Affected |',
        '|-----------|-------------------|',
    ]
    for flag, cnt in sorted(flag_freq.items(), key=lambda x: -x[1]):
        lines.append(f'| `{flag}` | {cnt} |')

    lines += [
        '',
        '---',
        '',
        '## Observation Quality Signal Frequency',
        '',
        'Top positive and negative signals across all observations:',
        '',
        '| Signal | Observations |',
        '|--------|--------------|',
    ]
    for sig, cnt in sorted(sig_freq.items(), key=lambda x: -x[1])[:20]:
        lines.append(f'| `{sig}` | {cnt} |')

    path = REPORTS_DIR / 'trust_distribution_summary.md'
    path.write_text('\n'.join(lines), encoding='utf-8')
    log.info('trust_distribution_summary.md → %s', path)


def write_batch_report_md(stats: dict, run_date: str) -> None:
    products = stats['products']

    # Barcode quality distribution
    bc_dist: dict[str, int] = {}
    for p in products:
        status = p.get('barcode_validation_status') or 'unknown'
        bc_dist[status] = bc_dist.get(status, 0) + 1

    # Nutrition consistency distribution
    nc_dist: dict[str, int] = {}
    for p in products:
        status = p.get('nutrition_consistency_status') or 'unknown'
        nc_dist[status] = nc_dist.get(status, 0) + 1

    # Ingredient quality distribution
    iq_dist: dict[str, int] = {}
    for p in products:
        quality = p.get('ingredient_text_quality') or 'unknown'
        iq_dist[quality] = iq_dist.get(quality, 0) + 1

    # Conflict category totals
    total_id_conf  = sum(p.get('identity_conflicts_count', 0)     for p in products)
    total_nu_conf  = sum(p.get('nutrition_conflicts_count', 0)    for p in products)
    total_ing_conf = sum(p.get('ingredient_conflicts_count', 0)   for p in products)
    total_lab_conf = sum(p.get('labeling_conflicts_count', 0)     for p in products)
    total_cmp_conf = sum(p.get('completeness_conflicts_count', 0) for p in products)

    lines = [
        f'# BSIP1 Batch Test 001 — Report',
        f'',
        f'**Run date:** {run_date}  ',
        f'**BSIP0 root:** `{BSIP0_ROOT}`  ',
        f'**Retailers:** {", ".join(RETAILER_IDS)}',
        f'',
        f'---',
        f'',
        f'## Summary',
        f'',
        f'| Metric | Count |',
        f'|--------|-------|',
        f'| Total observations | {stats["total_observations"]} |',
        f'| Barcode groups | {stats["barcode_groups"]} |',
        f'| → Multi-retailer groups | {stats["multi_retailer"]} |',
        f'| → Single-retailer groups | {stats["single_retailer"]} |',
        f'| No-barcode observations | {stats["no_barcode_obs"]} |',
        f'| Products written | {stats["products_written"]} |',
        f'| Schema valid | {stats["schema_valid"]} |',
        f'| Schema invalid | {stats["schema_invalid"]} |',
        f'',
        f'---',
        f'',
        f'## Barcode Quality Distribution',
        f'',
        f'| Status | Count |',
        f'|--------|-------|',
    ]
    for status, count in sorted(bc_dist.items(), key=lambda x: -x[1]):
        lines.append(f'| `{status}` | {count} |')

    lines += [
        f'',
        f'---',
        f'',
        f'## Nutrition Consistency Distribution',
        f'',
        f'| Status | Count |',
        f'|--------|-------|',
    ]
    for status, count in sorted(nc_dist.items(), key=lambda x: -x[1]):
        lines.append(f'| `{status}` | {count} |')

    lines += [
        f'',
        f'---',
        f'',
        f'## Ingredient Quality Distribution',
        f'',
        f'| Quality | Count |',
        f'|---------|-------|',
    ]
    for quality, count in sorted(iq_dist.items(), key=lambda x: -x[1]):
        lines.append(f'| `{quality}` | {count} |')

    lines += [
        f'',
        f'---',
        f'',
        f'## Conflict Category Totals (across all products)',
        f'',
        f'| Category | Total Conflict Records |',
        f'|----------|------------------------|',
        f'| Identity (name, brand, barcode, size) | {total_id_conf} |',
        f'| Nutrition (per-100g values) | {total_nu_conf} |',
        f'| Ingredient | {total_ing_conf} |',
        f'| Labeling (kosher, country, allergens) | {total_lab_conf} |',
        f'| Completeness (serving, image, etc.) | {total_cmp_conf} |',
        f'| **Total** | **{total_id_conf + total_nu_conf + total_ing_conf + total_lab_conf + total_cmp_conf}** |',
        f'',
        f'---',
        f'',
        f'## No-Barcode Candidates',
        f'',
    ]
    if stats['no_barcode_obs'] == 0:
        lines.append('None — all observations had a barcode.')
    else:
        lines.append(f'{stats["no_barcode_obs"]} observation(s) queued in `reports/fuzzy_candidate_queue.json`.')
        lines.append('')
        for p in stats.get('no_barcode_list', []):
            lines.append(
                f'- **{p["retailer_id"]}** — `{p["name"][:60]}` '
                f'(brand: {p["brand"] or "—"}, size: {p["size"] or "—"})'
            )

    lines += [
        f'',
        f'---',
        f'',
        f'## Products with Nutrition Warnings',
        f'',
    ]
    nutr_warn_products = [
        p for p in products
        if p.get('nutrition_consistency_warnings')
    ]
    if nutr_warn_products:
        lines.append(f'| Product | Status | Warnings |')
        lines.append(f'|---------|--------|----------|')
        for p in nutr_warn_products:
            warnings = p.get('nutrition_consistency_warnings', [])
            lines.append(
                f'| `{p["canonical_product_id"]}` '
                f'| `{p.get("nutrition_consistency_status", "")}` '
                f'| {len(warnings)} warning(s) |'
            )
    else:
        lines.append('No products with nutrition consistency warnings.')

    lines += [
        f'',
        f'---',
        f'',
        f'## Products with Ingredient Quality Issues',
        f'',
    ]
    ing_issue_products = [
        p for p in products
        if p.get('ingredient_text_quality') not in ('clean', 'missing', None)
    ]
    if ing_issue_products:
        lines.append(f'| Product | Quality | Warnings |')
        lines.append(f'|---------|---------|----------|')
        for p in ing_issue_products:
            lines.append(
                f'| `{p["canonical_product_id"]}` '
                f'| `{p.get("ingredient_text_quality", "")}` '
                f'| {", ".join(p.get("ingredient_warnings", [])[:1])!r} |'
            )
    else:
        lines.append('No products with ingredient quality issues.')

    report_path = REPORTS_DIR / 'batch_report.md'
    report_path.write_text('\n'.join(lines), encoding='utf-8')
    log.info('batch_report.md → %s', report_path)


# ── main ──────────────────────────────────────────────────────────────────────

def main() -> None:
    from datetime import datetime, timezone
    run_date = datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')

    log.info('=' * 64)
    log.info('BSIP1 batch_test_001 — start')
    log.info('BSIP0 root:  %s', BSIP0_ROOT)
    log.info('Output dir:  %s', OUTPUT_DIR)
    log.info('Reports dir: %s', REPORTS_DIR)
    log.info('openpyxl:    %s', 'available' if HAS_OPENPYXL else 'NOT INSTALLED — will write CSV')
    log.info('=' * 64)

    all_obs = read_bsip0_root(BSIP0_ROOT, RETAILER_IDS)
    barcode_groups, no_barcode = group_by_barcode(all_obs)

    stats = {
        'total_observations': len(all_obs),
        'barcode_groups':     len(barcode_groups),
        'no_barcode_obs':     len(no_barcode),
        'multi_retailer':     0,
        'single_retailer':    0,
        'products_written':   0,
        'schema_valid':       0,
        'schema_invalid':     0,
        'products':           [],
        'no_barcode_list':    [],
    }

    all_products: list[dict] = []
    all_obs_trust_rows: list[dict] = []

    for barcode, obs_list in sorted(barcode_groups.items()):
        retailer_set = {o['retailer_id'] for o in obs_list}
        is_multi     = len(retailer_set) > 1
        strategy     = 'multi_retailer_barcode_merge' if is_multi else 'single_retailer_passthrough'

        log.info('---')
        log.info('Barcode %s | %d obs | retailers: %s | strategy: %s',
                 barcode, len(obs_list), sorted(retailer_set), strategy)

        try:
            product, audit = canonicalize(obs_list, barcode)
        except Exception as exc:
            log.error('canonicalize failed for %s: %s', barcode, exc, exc_info=True)
            continue

        _write_pair(product, audit)

        prod_valid  = _validate(product, 'bsip1_product_schema_v0_1.json', f'{barcode}/product')
        audit_valid = _validate(audit,   'bsip1_audit_schema_v0_1.json',   f'{barcode}/audit')

        valid_ok = prod_valid == 'VALID' and audit_valid == 'VALID'
        if valid_ok:
            stats['schema_valid'] += 1
        else:
            stats['schema_invalid'] += 1

        if is_multi:
            stats['multi_retailer'] += 1
        else:
            stats['single_retailer'] += 1
        stats['products_written'] += 1

        # Collect per-observation trust rows from audit sidecar
        pid = product['canonical_product_id']
        for obs_rec in audit.get('retailer_observations', []):
            all_obs_trust_rows.append({
                'canonical_product_id':        pid,
                'barcode':                     barcode,
                'retailer_id':                 obs_rec['retailer_id'],
                'scrape_mode':                 obs_rec['scrape_mode'],
                'barcode_source':              obs_rec.get('barcode_source', ''),
                'nutrition_completeness':      obs_rec.get('nutrition_completeness', ''),
                'observation_quality_score':   obs_rec.get('observation_quality_score', 0.0),
                'observation_quality_level':   obs_rec.get('observation_quality_level', ''),
                'observation_quality_signals': obs_rec.get('observation_quality_signals', []),
            })

        log.info(
            '  → %s | trust=%s(%.2f) identity=%s barcode=%s(%s) nutrition=%s conflicts=%d | %s/%s',
            product['canonical_product_id'],
            product.get('canonical_trust_level', '?'),
            product.get('canonical_trust_score', 0.0),
            product['confidence']['identity_confidence'],
            product['confidence']['barcode_confidence'],
            product.get('barcode_validation_status', '?'),
            product.get('nutrition_consistency_status', '?'),
            product['conflicts_summary']['count'],
            prod_valid,
            audit_valid,
        )

        stats['products'].append({
            'canonical_product_id':       product['canonical_product_id'],
            'barcode':                    barcode,
            'strategy':                   strategy,
            'retailers':                  sorted(retailer_set),
            'observation_count':          len(obs_list),
            'identity_confidence':        product['confidence']['identity_confidence'],
            'barcode_confidence':         product['confidence']['barcode_confidence'],
            'nutrition_confidence':       product['confidence']['nutrition_confidence'],
            'conflicts':                  product['conflicts_summary']['count'],
            'inferred_fields':            product['inferred_fields'],
            'missing_fields':             product['missing_fields'],
            'product_valid':              prod_valid,
            'audit_valid':                audit_valid,
            # new validation fields
            'barcode_validation_status':      product.get('barcode_validation_status'),
            'barcode_confidence_reason':      product.get('barcode_confidence_reason'),
            'nutrition_consistency_status':   product.get('nutrition_consistency_status'),
            'nutrition_consistency_warnings': product.get('nutrition_consistency_warnings', []),
            'ingredient_text_quality':        product.get('ingredient_text_quality'),
            'ingredient_warnings':            product.get('ingredient_warnings', []),
            # conflict category counts for report aggregation
            'identity_conflicts_count':       len(product['conflicts_summary'].get('identity_conflicts', [])),
            'nutrition_conflicts_count':      len(product['conflicts_summary'].get('nutrition_conflicts', [])),
            'ingredient_conflicts_count':     len(product['conflicts_summary'].get('ingredient_conflicts', [])),
            'labeling_conflicts_count':       len(product['conflicts_summary'].get('labeling_conflicts', [])),
            'completeness_conflicts_count':   len(product['conflicts_summary'].get('completeness_conflicts', [])),
        })

        all_products.append(product)

    if no_barcode:
        log.info('---')
        log.info('%d no-barcode observations (queued for fuzzy match):', len(no_barcode))
        for o in no_barcode:
            name = (o.get('product_name_raw') or '')[:60]
            log.info('  %s  "%s"  brand=%s  size=%s',
                     o['retailer_id'], name, o.get('brand', '—'), o.get('package_size_raw', '—'))
            stats['no_barcode_list'].append({
                'retailer_id': o['retailer_id'],
                'name':        name,
                'brand':       o.get('brand'),
                'size':        o.get('package_size_raw'),
                'folder_path': o.get('folder_path', ''),
            })

    # Write run_summary.json (strip no_barcode_list — it's in the queue)
    summary = {k: v for k, v in stats.items() if k != 'no_barcode_list'}
    summary_path = OUTPUT_DIR / 'run_summary.json'
    summary_path.write_text(
        json.dumps(summary, ensure_ascii=False, indent=2), encoding='utf-8'
    )

    # Generate reports
    log.info('=' * 64)
    log.info('Generating reports → %s', REPORTS_DIR)

    write_barcode_quality_report(all_products)
    log.info('  barcode_quality_report')

    write_nutrition_warnings_report(all_products)
    log.info('  nutrition_consistency_warnings')

    write_ingredient_warnings_report(all_products)
    log.info('  ingredient_quality_warnings')

    write_conflict_breakdown_report(all_products)
    log.info('  conflict_breakdown')

    write_fuzzy_candidate_queue(no_barcode)

    write_observation_quality_report(all_obs_trust_rows)
    log.info('  observation_quality_report (%d rows)', len(all_obs_trust_rows))

    write_canonical_trust_report(all_products)
    log.info('  canonical_trust_report')

    write_high_risk_products_report(all_products)

    obs_count_stats = {
        'total':            stats['total_observations'],
        'scored':           len(all_obs_trust_rows),
        'excluded':         stats['no_barcode_obs'],
        'exclusion_reason': (
            'No-barcode observations are not canonicalized into products; '
            'they are queued in fuzzy_candidate_queue.json for manual matching '
            'and do not contribute to observation trust scoring.'
        ),
    }
    write_trust_distribution_md(all_products, all_obs_trust_rows, run_date, obs_count_stats)

    write_batch_report_md(stats, run_date)
    log.info('  batch_report.md')

    log.info('=' * 64)
    log.info('Done.')
    log.info('  Total observations:   %d', stats['total_observations'])
    log.info('  Barcode groups:       %d', stats['barcode_groups'])
    log.info('    multi-retailer:     %d', stats['multi_retailer'])
    log.info('    single-retailer:    %d', stats['single_retailer'])
    log.info('  No-barcode obs:       %d', stats['no_barcode_obs'])
    log.info('  Products written:     %d', stats['products_written'])
    log.info('  Schema valid/invalid: %d / %d', stats['schema_valid'], stats['schema_invalid'])
    log.info('  Summary:             %s', summary_path)
    log.info('=' * 64)


if __name__ == '__main__':
    main()
